"""בדיקות לארבע התוספות המהירות: חיפוש+ייצוא בתקלות מדווחות, מחיקה קבוצתית (admin בלבד),
הדגשת איחור ברשימה, והתחברות אחרונה בפאנל הניהול."""
from datetime import date, timedelta


def _login_admin(client):
    from app.models.user import User
    admin = User.query.filter_by(role="admin").first()
    client.post("/login", data={"username": admin.username, "password": "Admin@2026!"})
    return admin


# ---------- חיפוש + ייצוא בתקלות מדווחות ----------

def test_reported_issues_search(client, db_session):
    admin = _login_admin(client)
    client.post("/report", data={"title": "ברז נוזל בקומה 3", "description": ""})
    client.post("/report", data={"title": "מעלית תקועה", "description": ""})

    r = client.get("/reported_issues?search=ברז")
    body = r.get_data(as_text=True)
    assert "ברז נוזל" in body
    assert "מעלית תקועה" not in body


def test_reported_issues_export_excel_only_includes_public(client, db_session):
    from app.models.user import User
    from app.models.task import Task
    import openpyxl, io

    admin = _login_admin(client)
    client.post("/report", data={"title": "דיווח לייצוא", "description": ""})
    db_session.add(Task(title="משימה פנימית לייצוא", user_id=admin.id, assigned_to_id=admin.id, source="internal"))
    db_session.commit()

    r = client.get("/export/excel?source=public")
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    ws = wb.active
    titles = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert "דיווח לייצוא" in titles
    assert "משימה פנימית לייצוא" not in titles


def test_main_export_excludes_public_by_default(client, db_session):
    from app.models.user import User
    from app.models.task import Task
    import openpyxl, io

    admin = _login_admin(client)
    client.post("/report", data={"title": "דיווח שלא בייצוא הרגיל", "description": ""})
    db_session.add(Task(title="משימה רגילה בייצוא", user_id=admin.id, assigned_to_id=admin.id, source="internal"))
    db_session.commit()

    r = client.get("/export/excel")
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    ws = wb.active
    titles = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert "משימה רגילה בייצוא" in titles
    assert "דיווח שלא בייצוא הרגיל" not in titles


# ---------- מחיקה קבוצתית (admin בלבד) ----------

def test_admin_can_bulk_delete(client, db_session):
    from app.models.user import User
    from app.models.task import Task

    admin = _login_admin(client)
    t1 = Task(title="למחיקה קבוצתית 1", user_id=admin.id, assigned_to_id=admin.id)
    t2 = Task(title="למחיקה קבוצתית 2", user_id=admin.id, assigned_to_id=admin.id)
    db_session.add_all([t1, t2])
    db_session.commit()
    ids = [t1.id, t2.id]

    r = client.post("/bulk_delete_tasks", data={"task_ids": [str(i) for i in ids]})
    data = r.get_json()
    assert data["success"] is True
    assert data["deleted_count"] == 2
    assert Task.query.filter(Task.id.in_(ids)).count() == 0


def test_manager_cannot_bulk_delete(client, db_session):
    from app.models.user import User
    from app.models.task import Task

    mgr = User(username="nobulkdeletemgr", role="manager")
    mgr.set_password("x")
    db_session.add(mgr)
    db_session.commit()

    task = Task(title="שורדת מחיקה", user_id=mgr.id, assigned_to_id=mgr.id)
    db_session.add(task)
    db_session.commit()
    task_id = task.id

    client.post("/login", data={"username": "nobulkdeletemgr", "password": "x"})
    r = client.post("/bulk_delete_tasks", data={"task_ids": [str(task_id)]})
    assert r.status_code == 403
    assert Task.query.get(task_id) is not None


def test_employee_cannot_bulk_delete(client, db_session):
    from app.models.user import User

    emp = User(username="nobulkdeleteemp", role="employee")
    emp.set_password("x")
    db_session.add(emp)
    db_session.commit()

    client.post("/login", data={"username": "nobulkdeleteemp", "password": "x"})
    r = client.post("/bulk_delete_tasks", data={"task_ids": ["1"]})
    assert r.status_code == 403


def test_bulk_delete_creates_audit_entry(client, db_session):
    from app.models.user import User
    from app.models.task import Task
    from app.models.audit_log import AuditLog

    admin = _login_admin(client)
    task = Task(title="לביקורת מחיקה קבוצתית", user_id=admin.id, assigned_to_id=admin.id)
    db_session.add(task)
    db_session.commit()

    client.post("/bulk_delete_tasks", data={"task_ids": [str(task.id)]})

    entry = AuditLog.query.filter_by(action="bulk_delete_tasks").first()
    assert entry is not None
    assert "לביקורת מחיקה קבוצתית" in entry.details


def test_bulk_delete_button_hidden_for_manager(client, db_session):
    from app.models.user import User

    mgr = User(username="nodeletebtnmgr", role="manager")
    mgr.set_password("x")
    db_session.add(mgr)
    db_session.commit()

    client.post("/login", data={"username": "nodeletebtnmgr", "password": "x"})
    r = client.get("/")
    body = r.get_data(as_text=True)
    assert "מחק את הנבחרות" not in body


def test_bulk_delete_button_shown_for_admin(client, db_session):
    _login_admin(client)
    r = client.get("/")
    body = r.get_data(as_text=True)
    assert "מחק את הנבחרות" in body


# ---------- הדגשת איחור ברשימה ----------

def test_overdue_task_shows_badge_in_list(client, db_session):
    from app.models.task import Task

    admin = _login_admin(client)
    task = Task(title="משימה באיחור ברשימה", user_id=admin.id, assigned_to_id=admin.id,
                due_date=date.today() - timedelta(days=2), status="TODO")
    db_session.add(task)
    db_session.commit()

    r = client.get("/")
    body = r.get_data(as_text=True)
    assert "באיחור 2 ימים" in body


def test_non_overdue_task_no_badge(client, db_session):
    from app.models.task import Task

    admin = _login_admin(client)
    task = Task(title="משימה עתידית תקינה", user_id=admin.id, assigned_to_id=admin.id,
                due_date=date.today() + timedelta(days=2), status="TODO")
    db_session.add(task)
    db_session.commit()

    r = client.get("/")
    body = r.get_data(as_text=True)
    assert "באיחור" not in body


def test_done_overdue_task_no_badge(client, db_session):
    """משימה שהושלמה לא אמורה להיחשב 'באיחור' גם אם תאריך היעד עבר."""
    from app.models.task import Task

    admin = _login_admin(client)
    task = Task(title="משימה שהושלמה אחרי היעד", user_id=admin.id, assigned_to_id=admin.id,
                due_date=date.today() - timedelta(days=2), status="DONE")
    db_session.add(task)
    db_session.commit()

    r = client.get("/")
    body = r.get_data(as_text=True)
    assert "באיחור" not in body


# ---------- התחברות אחרונה ----------

def test_last_login_recorded_on_login(client, db_session):
    from app.models.user import User

    admin = _login_admin(client)
    refreshed = User.query.filter_by(username=admin.username).first()
    assert refreshed.last_login_at is not None


def test_last_login_shown_in_admin_panel(client, db_session):
    from app.models.user import User

    emp = User(username="neverloggedinuser", role="employee")
    emp.set_password("x")
    db_session.add(emp)
    db_session.commit()

    _login_admin(client)
    r = client.get("/admin")
    body = r.get_data(as_text=True)
    assert "מעולם לא התחבר" in body


def test_last_login_updates_on_each_login(client, db_session, monkeypatch):
    import time
    from app.models.user import User

    emp = User(username="loginupdateuser", role="employee")
    emp.set_password("x")
    db_session.add(emp)
    db_session.commit()

    client.post("/login", data={"username": "loginupdateuser", "password": "x"})
    first = User.query.filter_by(username="loginupdateuser").first().last_login_at
    assert first is not None

    client.get("/logout")
    time.sleep(0.05)
    client.post("/login", data={"username": "loginupdateuser", "password": "x"})
    second = User.query.filter_by(username="loginupdateuser").first().last_login_at
    assert second >= first
