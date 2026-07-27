from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app, send_file, session
import os
import base64
import io
from flask_login import login_user, logout_user, login_required, current_user
from app.models.task import Task, TaskChecklistItem, RECURRENCE_NONE, RECURRENCE_DAILY, RECURRENCE_WEEKLY, RECURRENCE_MONTHLY, MAX_IMAGE_SIZE_BYTES
from app.models.user import User, validate_password_strength
from app.models.department import Department
from app.models.comment import TaskComment
from app.models.notification import Notification, notify, notify_with_email, queue_email, notify_recipients_multi_channel
from app.models.task_template import TaskTemplate
from app.models.audit_log import log_audit
from app.services import automation_service
from app import db, mail, limiter
from flask_mail import Message
from datetime import datetime, date, timedelta
from sqlalchemy import text

bp = Blueprint("tasks", __name__)

# =========================================================
# System Health Check (Phase 4)
# =========================================================
@bp.route("/api/health")
def health_check():
    """מספק נקודת חיבור (Readiness/Liveness Probe) לבדיקת זמינות מסד הנתונים"""
    try:
        db.session.execute(text('SELECT 1'))
        db_status = True
    except Exception as e:
        db_status = False
    
    return jsonify({
        "status": "ok" if db_status else "error",
        "db_online": db_status,
        "timestamp": datetime.utcnow().isoformat()
    }), 200 if db_status else 500


@bp.route("/service-worker.js")
def service_worker():
    """
    מגישים את קובץ ה-Service Worker מנתיב השורש (לא /static/) כדי שההיקף (scope) שלו
    יכסה את כל האתר - זה תנאי הכרחי לכך שה-PWA יהיה ניתן להתקנה על כל עמוד, לא רק תחת /static/.
    """
    path = os.path.join(current_app.root_path, "static", "service-worker-src.js")
    with open(path, "r", encoding="utf-8") as f:
        content = f.read()
    response = current_app.response_class(content, mimetype="application/javascript")
    response.headers["Service-Worker-Allowed"] = "/"
    return response


def can_create_tasks(user):
    return user.role in ("admin", "manager")


def visible_task_query(user):
    """שאילתת המשימות שהמשתמש רשאי לראות, בהתאם לתפקיד ולמחלקה."""
    if user.role == "admin":
        return Task.query
    if user.role == "manager":
        return Task.query.filter(Task.assigned_to_id.in_(user.visible_user_ids()))
    return Task.query.filter_by(assigned_to_id=user.id)


def apply_task_filters(query, user, args):
    """מחיל את כל הסינונים (חיפוש/סטטוס/עדיפות/אחראי/טווח תאריכים/חזרתיות) על שאילתת משימות.
    משותף בין תצוגת הרשימה הרגילה לבין ייצוא הדוחות, כדי שהדוח תמיד יתאים בדיוק למה שרואים על המסך."""
    search_query = args.get("search", "")
    if search_query:
        query = query.filter((Task.title.contains(search_query)) | (Task.description.contains(search_query)))

    status_filter = args.get("status", "")
    if status_filter:
        query = query.filter(Task.status == status_filter)

    priority_filter = args.get("priority", "")
    if priority_filter:
        query = query.filter(Task.priority == priority_filter)

    assignee_filter = args.get("assignee", "")
    if assignee_filter and assignee_filter.isdigit() and int(assignee_filter) in user.visible_user_ids():
        query = query.filter(Task.assigned_to_id == int(assignee_filter))

    due_after = args.get("due_after", "")
    if due_after:
        try:
            query = query.filter(Task.due_date >= datetime.strptime(due_after, "%Y-%m-%d").date())
        except ValueError:
            pass
    due_before = args.get("due_before", "")
    if due_before:
        try:
            query = query.filter(Task.due_date <= datetime.strptime(due_before, "%Y-%m-%d").date())
        except ValueError:
            pass

    recurring_only = args.get("recurring_only", "")
    if recurring_only == "1":
        query = query.filter(Task.recurrence.isnot(None), Task.recurrence != "NONE")

    # ברירת מחדל: לא לכלול דיווחים ציבוריים (יש להם לשונית/ייצוא נפרדים) -
    # אלא אם ביקשו אותם במפורש (source=public, למשל מייצוא בתוך לשונית "תקלות מדווחות")
    source_filter = args.get("source", "")
    if source_filter == "public":
        query = query.filter(Task.source == "public")
    elif source_filter != "all":
        query = query.filter(Task.source != "public")

    sort_by = args.get("sort", "created_at")
    if sort_by == "due_date":
        query = query.order_by(Task.due_date.asc())
    elif sort_by == "priority":
        priority_order = db.case((Task.priority == "HIGH", 0), (Task.priority == "MEDIUM", 1), else_=2)
        query = query.order_by(priority_order)
    else:
        query = query.order_by(Task.created_at.desc())

    return query


def can_touch_task(user, task):
    """האם המשתמש רשאי לערוך/למחוק/לשנות סטטוס של משימה ספציפית."""
    if user.role == "admin":
        return True
    if user.role == "manager":
        return task.assigned_to_id in user.visible_user_ids()
    return task.assigned_to_id == user.id


def _next_due_date(current_due_date, recurrence):
    """מחשב את תאריך היעד הבא עבור משימה חוזרת."""
    base = current_due_date or date.today()
    if recurrence == RECURRENCE_DAILY:
        return base + timedelta(days=1)
    if recurrence == RECURRENCE_WEEKLY:
        return base + timedelta(weeks=1)
    if recurrence == RECURRENCE_MONTHLY:
        month = base.month + 1
        year = base.year + (1 if month > 12 else 0)
        month = 1 if month > 12 else month
        day = base.day
        while True:
            try:
                return base.replace(year=year, month=month, day=day)
            except ValueError:
                day -= 1
    return base


def _create_next_recurrence(task):
    """כשמשימה חוזרת מסומנת כ-DONE, יוצרים אוטומטית את המופע הבא שלה."""
    if task.recurrence == RECURRENCE_NONE or not task.recurrence:
        return None

    next_task = Task(
        title=task.title,
        description=task.description,
        priority=task.priority,
        status="TODO",
        user_id=task.user_id,
        assigned_to_id=task.assigned_to_id,
        due_date=_next_due_date(task.due_date, task.recurrence),
        recurrence=task.recurrence,
        recurrence_parent_id=task.id,
    )
    db.session.add(next_task)
    db.session.commit()
    return next_task

# =========================================================
# 🔒 מערכת ניהול משתמשים (Authentication)
# =========================================================

@bp.route("/login", methods=["GET", "POST"])
@limiter.limit("10 per 15 minutes", methods=["POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("tasks.index"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password")
        remember_me = request.form.get("remember_me") == "1"
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user, remember=remember_me)
            session.permanent = True  # מפעיל את PERMANENT_SESSION_LIFETIME גם בלי "זכור אותי"
            user.last_login_at = datetime.utcnow()
            db.session.commit()
            flash(f"ברוך הבא, {user.username}!", "success")
            return redirect(url_for("tasks.index"))
        else:
            flash("שם משתמש או סיסמה לא נכונים.", "danger")
    return render_template("login.html")

@bp.route("/register", methods=["GET", "POST"])
@limiter.limit("5 per hour", methods=["POST"])
def register():
    if current_user.is_authenticated:
        return redirect(url_for("tasks.index"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip() or None
        phone = request.form.get("phone", "").strip() or None
        password = request.form.get("password")

        if not username:
            flash("יש להזין שם משתמש.", "danger")
            return redirect(url_for("tasks.register"))

        is_strong, error_msg = validate_password_strength(password)
        if not is_strong:
            flash(error_msg, "danger")
            return redirect(url_for("tasks.register"))

        existing = User.query.filter_by(username=username).first()
        if existing or (email and User.query.filter_by(email=email).first()):
            flash("שם המשתמש או האימייל כבר קיימים.", "danger")
            return redirect(url_for("tasks.register"))
        try:
            user = User(username=username, email=email, phone=phone)
            user.set_password(password)
            db.session.add(user)
            db.session.commit()
            login_user(user, remember=True)
            session.permanent = True
            flash("החשבון נוצר בהצלחה! 🎉", "success")
            return redirect(url_for("tasks.index"))
        except Exception as e:
            db.session.rollback()
            flash(f"שגיאה ברישום: {e}", "danger")
            return redirect(url_for("tasks.register"))
    return render_template("register.html")


@bp.route("/report", methods=["GET", "POST"])
@limiter.limit("8 per hour")
def public_report():
    departments = Department.query.order_by(Department.name).all()

    if request.method == "POST":
        # honeypot אנטי-ספאם: שדה מוסתר ב-CSS שבני אדם לעולם לא ימלאו, רק בוטים אוטומטיים
        if request.form.get("website"):
            flash("הדיווח נשלח בהצלחה, תודה!", "success")
            return redirect(url_for("tasks.public_report"))

        title = request.form.get("title", "").strip()
        description = request.form.get("description", "").strip()
        department_id = request.form.get("department_id") or None
        reporter_name = request.form.get("reporter_name", "").strip() or None
        reporter_phone = request.form.get("reporter_phone", "").strip() or None

        if not title:
            flash("יש לתאר בקצרה מהי התקלה.", "danger")
            return render_template("report.html", departments=departments)

        department = Department.query.get(int(department_id)) if department_id else None

        # מנסים לשייך משימה למנהל התחום הרלוונטי; אם אין כזה, נופלים חזרה לכל מנהלי המערכת
        assigned_to_id = None
        recipients = []
        if department:
            dept_managers = User.query.filter_by(department_id=department.id, role="manager").all()
            recipients = dept_managers
            if dept_managers:
                assigned_to_id = dept_managers[0].id
        if not recipients:
            recipients = User.query.filter_by(role="admin").all()
            if not assigned_to_id and recipients:
                assigned_to_id = recipients[0].id

        task = Task(
            title=title[:100],
            description=description,
            priority="MEDIUM",
            status="TODO",
            department_id=department.id if department else None,
            assigned_to_id=assigned_to_id,
            source="public",
            reporter_name=reporter_name,
            reporter_phone=reporter_phone,
        )
        db.session.add(task)
        db.session.commit()

        for recipient in recipients:
            notify_with_email(
                recipient,
                f"דווחה תקלה חדשה: \"{task.title}\"" + (f" (דיווח מ-{reporter_name})" if reporter_name else ""),
                link=f"/edit/{task.id}",
                icon="bi-megaphone",
                email_subject=f"דיווח תקלה חדש: {task.title}",
                email_body=(
                    f"דווחה תקלה חדשה דרך הטופס הציבורי.\n\n"
                    f"כותרת: {task.title}\n"
                    f"{'תיאור: ' + description if description else ''}\n"
                    f"{'מחלקה: ' + department.name if department else ''}\n"
                    f"{'שם המדווח: ' + reporter_name if reporter_name else 'המדווח לא השאיר שם'}\n"
                    f"{'טלפון לחזרה: ' + reporter_phone if reporter_phone else ''}"
                )
            )

        automation_service.trigger("report_submitted", task)

        flash("הדיווח נשלח בהצלחה! הצוות הרלוונטי קיבל התראה. תודה שדיווחת!", "success")
        return redirect(url_for("tasks.public_report"))

    return render_template("report.html", departments=departments)


@bp.route("/logout")
@login_required
def logout():
    logout_user()
    flash("התנתקת בהצלחה.", "success")
    return redirect(url_for("tasks.login"))


@bp.route("/change_password", methods=["GET", "POST"])
@login_required
def change_password():
    """
    מסך שינוי סיסמה חובה - מוצג אוטומטית (ר' before_request ב-app/__init__.py) כשלמשתמש
    יש must_change_password=True, למשל אחרי איפוס דרך /rescue לסיסמה ידועה.
    """
    if request.method == "POST":
        new_password = request.form.get("password")
        is_strong, error_msg = validate_password_strength(new_password)
        if not is_strong:
            flash(error_msg, "danger")
            return render_template("change_password.html")

        current_user.set_password(new_password)
        current_user.must_change_password = False
        db.session.commit()
        flash("הסיסמה הוחלפה בהצלחה!", "success")
        return redirect(url_for("tasks.index"))

    return render_template("change_password.html")


# =========================================================
# 📋 ניהול משימות והרשאות (Core Task & Roles)
# =========================================================

@bp.route("/", methods=["GET", "POST"])
@login_required
def index():
    # שליפת המשתמשים הרלוונטיים עבור טופס ההקצאה (לפי היקף ההרשאה של המשתמש)
    all_users = current_user.visible_users_query().order_by(User.username).all() if can_create_tasks(current_user) else []

    if request.method == "POST":
        if not can_create_tasks(current_user):
            flash("רק מנהלים מורשים ליצור משימות חדשות.", "danger")
            return redirect(url_for("tasks.index"))

        due_date_str = request.form.get("due_date")
        due_date = None
        if due_date_str:
            try:
                parsed_date = datetime.strptime(due_date_str, "%Y-%m-%d").date()
                if parsed_date.year <= 9999:
                    due_date = parsed_date
            except ValueError:
                pass

        assigned_to_id = request.form.get("assigned_to_id")
        if not assigned_to_id:
            assigned_to_id = current_user.id
        else:
            assigned_to_id = int(assigned_to_id)
            # מנהל תחום לא יכול להקצות משימה למישהו מחוץ למחלקה שלו
            if assigned_to_id not in current_user.visible_user_ids():
                flash("אינך רשאי להקצות משימה למשתמש זה.", "danger")
                return redirect(url_for("tasks.index"))

        new_recurrence = request.form.get("recurrence", "NONE")
        if new_recurrence not in ("NONE", "DAILY", "WEEKLY", "MONTHLY"):
            new_recurrence = "NONE"

        task = Task(
            title=request.form.get("title", ""),
            description=request.form.get("description", ""),
            due_date=due_date,
            priority=request.form.get("priority", "LOW"),
            user_id=current_user.id,
            assigned_to_id=assigned_to_id,
            recurrence=new_recurrence,
        )
        db.session.add(task)
        db.session.commit()

        if task.assigned_to_id and task.assigned_to_id != current_user.id:
            assignee_obj = User.query.get(task.assigned_to_id)
            email_ok = notify_with_email(
                assignee_obj,
                f"{current_user.username} הקצה לך משימה חדשה: \"{task.title}\"",
                link=f"/edit/{task.id}",
                icon="bi-plus-circle",
                email_subject=f"משימה חדשה הוקצתה לך: {task.title}",
                email_body=(
                    f"שלום {assignee_obj.username},\n\n"
                    f"{current_user.username} הקצה לך משימה חדשה:\n\n"
                    f"כותרת: {task.title}\n"
                    f"{'תיאור: ' + task.description if task.description else ''}\n"
                    f"{'תאריך יעד: ' + task.due_date.strftime('%d/%m/%Y') if task.due_date else ''}\n"
                    f"עדיפות: {_PRIORITY_LABELS_HE.get(task.priority, task.priority)}"
                )
            )
            if assignee_obj.email and not email_ok:
                flash(f"⚠️ המשימה נוצרה, אבל שליחת המייל ל-{assignee_obj.username} נכשלה - בדוק את הגדרות SMTP.", "danger")

        automation_service.trigger("task_created", task)

        flash("המשימה נוצרה והוקצתה בהצלחה!", "success")
        return redirect(url_for("tasks.index"))

    # לוגיקת סינון: מנהל מערכת רואה הכל, מנהל תחום רואה את מחלקתו, עובד רואה רק את עצמו
    # דיווחים ציבוריים לא מוצגים כאן בכלל (ברירת המחדל של apply_task_filters) - יש להם
    # לשונית נפרדת (/reported_issues), לפי בקשה מפורשת שלא לערבב בין דיווחים למשימות רגילות.
    query = visible_task_query(current_user)
    query = apply_task_filters(query, current_user, request.args)

    page = request.args.get("page", 1, type=int)
    sort_by = request.args.get("sort", "created_at")

    pagination = db.paginate(query, page=page, per_page=20, error_out=False)

    # תבניות רלוונטיות למודל יצירת משימה מהירה
    templates = []
    if can_create_tasks(current_user):
        if current_user.role == "admin":
            templates = TaskTemplate.query.order_by(TaskTemplate.name).all()
        else:
            templates = TaskTemplate.query.filter(
                (TaskTemplate.department_id == current_user.department_id) | (TaskTemplate.department_id.is_(None))
            ).order_by(TaskTemplate.name).all()

    # מחלקות זמינות לשליחת הודעה קבוצתית - admin רואה הכל, מנהל תחום רק את שלו
    messaging_departments = []
    if current_user.role == "admin":
        messaging_departments = Department.query.order_by(Department.name).all()
    elif current_user.role == "manager" and current_user.department:
        messaging_departments = [current_user.department]

    return render_template(
        "tasks.html", 
        tasks=pagination.items, 
        pagination=pagination, 
        sort_by=sort_by, 
        all_users=all_users, 
        now_date=date.today(),
        templates=templates,
        messaging_departments=messaging_departments
    )

@bp.route("/done/<int:id>")
@login_required
def done(id):
    task = db.session.get(Task, id)
    if not task:
        flash("המשימה לא נמצאה.", "danger")
        return redirect(url_for("tasks.index"))
        
    if not can_touch_task(current_user, task):
        flash("אינך רשאי לעדכן משימה זו.", "danger")
        return redirect(url_for("tasks.index"))
        
    # בדיקת תלויות חסרות (Soft Block)
    uncompleted_deps = [dep for dep in task.dependencies if dep.status != "DONE"]
    if uncompleted_deps and not request.args.get("force"):
        flash(f"לא ניתן לסיים. משימה זו תלויה במשימות שטרם הושלמו.", "warning")
        return redirect(url_for("tasks.index"))

    was_done_already = task.status == "DONE"
    task.status = "DONE"
    db.session.commit()
    next_task = _create_next_recurrence(task)
    if not was_done_already:
        automation_service.trigger("status_changed", task)
    if next_task:
        flash(f"כל הכבוד! המשימה בוצעה 🎉 (נוצרה אוטומטית משימה חוזרת חדשה לתאריך {next_task.due_date.strftime('%d/%m/%Y')})", "success")
    else:
        flash("כל הכבוד! המשימה בוצעה 🎉", "success")
    return redirect(url_for("tasks.index"))

@bp.route("/delete/<int:id>", methods=["POST"])
@login_required
def delete(id):
    task = db.session.get(Task, id)

    def _redirect_back():
        # חוזרים לעמוד שממנו הגיעה בקשת המחיקה (למשל לשונית "תקלות מדווחות"),
        # ולא תמיד לרשימת המשימות הרגילה - רק אם זה עמוד בתוך האתר שלנו עצמו.
        ref = request.referrer
        if ref and ref.startswith(request.host_url):
            return redirect(ref)
        return redirect(url_for("tasks.index"))

    if not task or not can_create_tasks(current_user) or not can_touch_task(current_user, task):
        flash("אין לך הרשאה למחוק משימה זו.", "danger")
        return _redirect_back()

    task_title = task.title
    db.session.delete(task)
    db.session.commit()
    log_audit(current_user, "delete_task", target_type="task", target_id=id, target_label=task_title)
    flash("המשימה נמחקה בהצלחה.", "danger")
    return _redirect_back()

@bp.route("/edit/<int:id>", methods=["GET", "POST"])
@login_required
def edit(id):
    task = db.session.get(Task, id)
    if not task:
        flash("המשימה לא נמצאה.", "danger")
        return redirect(url_for("tasks.index"))
        
    if not can_create_tasks(current_user) or not can_touch_task(current_user, task):
        flash("אין לך הרשאה לערוך משימה זו. צפייה בלבד.", "danger")
        return redirect(url_for("tasks.index"))

    if request.method == "POST":
        task.title = request.form.get("title", "")
        task.description = request.form.get("description", "")
        task.priority = request.form.get("priority", "LOW")

        new_recurrence = request.form.get("recurrence", "NONE")
        if new_recurrence in ("NONE", "DAILY", "WEEKLY", "MONTHLY"):
            task.recurrence = new_recurrence

        new_status = request.form.get("status")
        was_done_already = task.status == "DONE"
        old_status = task.status
        if new_status in ("TODO", "IN_PROGRESS", "DONE"):
            task.status = new_status

        new_assignee = request.form.get("assigned_to_id")
        previous_assignee_id = task.assigned_to_id
        if new_assignee:
            new_assignee = int(new_assignee)
            # ההקצאה החדשה חייבת להיות לתוך היקף ההרשאה של המשתמש (מחלקתו, או admin לכל אחד)
            if new_assignee in current_user.visible_user_ids():
                task.assigned_to_id = new_assignee

        due_date_str = request.form.get("due_date")
        if due_date_str:
            try:
                task.due_date = datetime.strptime(due_date_str, "%Y-%m-%d").date()
            except ValueError:
                pass
        else:
            task.due_date = None

        # Phase 3 Fields
        task.estimated_minutes = int(request.form.get("estimated_minutes") or 0)
        
        parent_task_id = request.form.get("parent_task_id")
        if parent_task_id and parent_task_id.isdigit():
            task.parent_task_id = int(parent_task_id)
        else:
            task.parent_task_id = None
            
        dependency_ids = request.form.getlist("dependency_ids")
        task.dependencies = []
        if dependency_ids:
            deps = Task.query.filter(Task.id.in_(dependency_ids)).all()
            for dep in deps:
                if dep.id != task.id: # Prevent self-dependency
                    task.dependencies.append(dep)

        # תמונה מצורפת (אופציונלי) - נשמרת כ-base64 ישירות ב-DB
        uploaded_file = request.files.get("image")
        if uploaded_file and uploaded_file.filename:
            file_bytes = uploaded_file.read()
            if len(file_bytes) > MAX_IMAGE_SIZE_BYTES:
                flash("התמונה גדולה מדי (מקסימום 2MB). שאר השינויים נשמרו, התמונה לא הועלתה.", "danger")
            elif uploaded_file.mimetype not in ("image/jpeg", "image/png", "image/webp", "image/gif"):
                flash("סוג קובץ לא נתמך (רק JPEG/PNG/WEBP/GIF). שאר השינויים נשמרו, התמונה לא הועלתה.", "danger")
            else:
                task.image_data = base64.b64encode(file_bytes).decode("ascii")
                task.image_mimetype = uploaded_file.mimetype

        if request.form.get("remove_image") == "1":
            task.image_data = None
            task.image_mimetype = None

        db.session.commit()

        if task.status != old_status:
            automation_service.trigger("status_changed", task)

        if task.status == "DONE" and not was_done_already:
            _create_next_recurrence(task)
            if task.user_id and task.user_id != current_user.id:
                notify(task.user_id, f"{current_user.username} סימן כבוצע: \"{task.title}\"",
                       link=f"/edit/{task.id}", icon="bi-check-circle")

        if task.assigned_to_id != previous_assignee_id and task.assigned_to_id and task.assigned_to_id != current_user.id:
            new_assignee_obj = User.query.get(task.assigned_to_id)
            email_ok = notify_with_email(
                new_assignee_obj,
                f"{current_user.username} הקצה לך משימה: \"{task.title}\"",
                link=f"/edit/{task.id}", icon="bi-arrow-left-right",
                email_subject=f"הוקצתה לך משימה: {task.title}",
                email_body=(
                    f"שלום {new_assignee_obj.username},\n\n"
                    f"{current_user.username} שייך אליך את המשימה \"{task.title}\".\n"
                    f"{'תאריך יעד: ' + task.due_date.strftime('%d/%m/%Y') if task.due_date else ''}"
                )
            )
            if new_assignee_obj.email and not email_ok:
                flash(f"⚠️ המשימה עודכנה, אבל שליחת המייל ל-{new_assignee_obj.username} נכשלה - בדוק את הגדרות SMTP.", "danger")

        flash("השינויים נשמרו והמשימה עודכנה.", "success")
        return redirect(url_for("tasks.index"))

    assignable_users = current_user.visible_users_query().order_by(User.username).all()
    comments = task.comments.order_by(TaskComment.created_at.asc()).all()
    all_other_tasks = visible_task_query(current_user).filter(Task.id != task.id).all()
    return render_template("edit_task.html", task=task, assignable_users=assignable_users, comments=comments, all_tasks=all_other_tasks)


@bp.route("/task/<int:id>/comment", methods=["POST"])
@login_required
def add_comment(id):
    task = db.session.get(Task, id)
    if not task or not can_touch_task(current_user, task):
        flash("אינך רשאי להגיב על משימה זו.", "danger")
        return redirect(url_for("tasks.index"))

    body = request.form.get("body", "").strip()
    if body:
        comment = TaskComment(task_id=task.id, user_id=current_user.id, body=body)
        db.session.add(comment)
        db.session.commit()

        recipients = {task.assigned_to_id, task.user_id} - {current_user.id, None}
        for recipient_id in recipients:
            notify(recipient_id, f"{current_user.username} הגיב על \"{task.title}\"",
                   link=f"/edit/{task.id}", icon="bi-chat-left-text")

        flash("התגובה נוספה.", "success")

    return redirect(url_for("tasks.edit", id=task.id))


# =========================================================
# Phase 3: Checklists
# =========================================================

@bp.route("/task/<int:task_id>/checklist", methods=["POST"])
@login_required
def add_checklist_item(task_id):
    task = db.session.get(Task, task_id)
    if not task or not can_touch_task(current_user, task):
        return jsonify({"error": "Unauthorized"}), 403
        
    data = request.json
    text = data.get('text')
    
    if text:
        item = TaskChecklistItem(task_id=task.id, text=text)
        db.session.add(item)
        db.session.commit()
        return jsonify({"success": True, "item_id": item.id, "text": item.text})
    return jsonify({"error": "Invalid data"}), 400

@bp.route("/checklist/<int:item_id>/toggle", methods=["POST"])
@login_required
def toggle_checklist_item(item_id):
    item = db.session.get(TaskChecklistItem, item_id)
    if not item or not can_touch_task(current_user, item.task):
        return jsonify({"error": "Unauthorized"}), 403
        
    item.is_done = not item.is_done
    db.session.commit()
    return jsonify({"success": True, "is_done": item.is_done})


# =========================================================
# Phase 3: Timeline
# =========================================================

@bp.route("/timeline")
@login_required
def timeline():
    tasks = visible_task_query(current_user).filter(Task.due_date.isnot(None), Task.status != 'DONE', Task.source != 'public').order_by(Task.created_at).all()
    return render_template('timeline.html', tasks=tasks)


# =========================================================
# 🚀 תצוגות מתקדמות (Kanban & Calendar)
# =========================================================

@bp.route("/reported_issues")
@login_required
def reported_issues():
    """
    לשונית עצמאית לחלוטין לדיווחים ציבוריים - במכוון לא מעורבבת עם רשימת המשימות הרגילה.
    זמין רק ל-admin/manager (עובד רגיל לא אמור לטפל בדיווחי ציבור).
    """
    if not can_create_tasks(current_user):
        flash("אין לך הרשאה לצפות בדיווחים ציבוריים.", "danger")
        return redirect(url_for("tasks.index"))

    # מכריחים source=public בלי קשר למה שהגיע מהטופס - זו לשונית ייעודית לדיווחים בלבד
    forced_args = request.args.copy()
    forced_args = forced_args.to_dict()
    forced_args["source"] = "public"

    query = visible_task_query(current_user)
    query = apply_task_filters(query, current_user, forced_args)

    page = request.args.get("page", 1, type=int)
    pagination = db.paginate(query, page=page, per_page=20, error_out=False)

    report_url = url_for("tasks.public_report", _external=True)

    return render_template(
        "reported_issues.html",
        issues=pagination.items,
        pagination=pagination,
        report_url=report_url,
    )


@bp.route("/kanban")
@login_required
def kanban():
    # דיווחים ציבוריים לא מוצגים כאן (יש להם לשונית נפרדת) - עקביות עם הרשימה הראשית
    base_query = visible_task_query(current_user).filter(Task.source != "public")

    todo = base_query.filter(Task.status.in_(["TODO", None])).order_by(Task.created_at.desc()).all()
    in_progress = base_query.filter(Task.status == "IN_PROGRESS").order_by(Task.created_at.desc()).all()
    # עמודת "בוצע" מוגבלת ל-50 האחרונות - עם הזמן היא רק גדלה ואין צורך להציג הכל
    done = base_query.filter(Task.status == "DONE").order_by(Task.created_at.desc()).limit(50).all()

    return render_template("kanban.html", todo=todo, in_progress=in_progress, done=done)


@bp.route("/update_status/<int:id>", methods=["POST"])
@login_required
def update_status(id):
    task = db.session.get(Task, id)
    if task and not can_touch_task(current_user, task):
        task = None

    if task:
        data = request.get_json()
        new_status = data.get("status")
        force_complete = data.get("force_complete", False)
        
        if new_status in ["TODO", "IN_PROGRESS", "DONE"]:
            # Soft Dependency Block
            if new_status == 'DONE' and not force_complete:
                uncompleted_deps = [dep.title for dep in task.dependencies if dep.status != 'DONE']
                if uncompleted_deps:
                    return jsonify({
                        "success": False,
                        "warning": True, 
                        "message": f"תלויות שטרם הושלמו: {', '.join(uncompleted_deps)}. להשלים למרות זאת?",
                        "uncompleted_deps": uncompleted_deps
                    }), 200

            was_done_already = task.status == "DONE"
            status_changed = new_status != task.status
            task.status = new_status
            db.session.commit()
            if new_status == "DONE" and not was_done_already:
                _create_next_recurrence(task)
            if status_changed:
                automation_service.trigger("status_changed", task)
            return jsonify({"success": True})
    return jsonify({"success": False}), 400


@bp.route("/calendar")
@login_required
def calendar():
    return render_template("calendar.html")

@bp.route("/api/calendar_tasks")
@login_required
def calendar_tasks():
    # דיווחים ציבוריים לא מוצגים כאן (יש להם לשונית נפרדת) - עקביות עם הרשימה הראשית
    tasks = visible_task_query(current_user).filter(Task.due_date.isnot(None), Task.source != "public").all()

    events = []
    for t in tasks:
        color = "#22c55e" if t.status == "DONE" else ("#ef4444" if t.priority == "HIGH" else "#3b82f6")
        can_edit = can_create_tasks(current_user) and can_touch_task(current_user, t)
        events.append({ 
            "id": t.id, 
            "title": t.title, 
            "start": t.due_date.isoformat(), 
            "backgroundColor": color, 
            "borderColor": color, 
            "url": f"/edit/{t.id}" if can_edit else "#" 
        })
    return jsonify(events)


@bp.route("/admin/test_email", methods=["POST"])
@login_required
@limiter.limit("10 per hour")
def test_email():
    """
    כלי אבחון ל-admin: שולח מייל בדיקה לכתובת שלו-עצמו ומחזיר תוצאה מדויקת -
    כדי לוודא שהגדרות MAIL_* ב-Render באמת עובדות, בלי לחפש בלוגים.
    כל שלב מתועד בנפרד ב-log (בלי לחשוף סיסמה/ערכים מלאים), וכל סוג כשל מוחזר
    עם הודעה שימושית וממוקדת במקום הודעה גנרית.
    """
    current_app.logger.info(f"[mail-test] בדיקת מייל התחילה - user_id={current_user.id}")

    if current_user.role != "admin":
        current_app.logger.warning(f"[mail-test] גישה נדחתה - user_id={current_user.id} אינו admin")
        return jsonify({"success": False, "message": "רק מנהל מערכת יכול להריץ בדיקה זו."}), 403

    # --- שלב 1: טעינת משתמש ---
    try:
        user = User.query.get(current_user.id)
    except Exception:
        current_app.logger.exception(f"[mail-test] שלב טעינת משתמש נכשל - user_id={current_user.id}")
        return jsonify({"success": False, "message": "מסד נתונים לא זמין - נסה שוב בעוד רגע."}), 500

    if not user:
        current_app.logger.error(f"[mail-test] שלב טעינת משתמש - user_id={current_user.id} לא נמצא ב-DB")
        return jsonify({"success": False, "message": "מסד נתונים לא זמין - המשתמש לא נמצא."}), 500

    current_app.logger.info(f"[mail-test] שלב טעינת משתמש הושלם - user_id={user.id}")

    if not user.email:
        current_app.logger.info(f"[mail-test] אין כתובת מייל מוגדרת - user_id={user.id}")
        return jsonify({"success": False, "message": "אין לך כתובת מייל מוגדרת בפרופיל שלך - הוסף אחת קודם."}), 400

    # --- בדיקת הגדרות SMTP לפני ניסיון שליחה (log בלבד, בלי לחשוף סיסמה/ערכים) ---
    mail_server = current_app.config.get("MAIL_SERVER")
    mail_port = current_app.config.get("MAIL_PORT")
    mail_username = current_app.config.get("MAIL_USERNAME")
    mail_password = current_app.config.get("MAIL_PASSWORD")

    current_app.logger.info(
        "[mail-test] בדיקת הגדרות SMTP: "
        f"MAIL_SERVER={'קיים' if mail_server else 'חסר'}, "
        f"MAIL_PORT={'קיים' if mail_port else 'חסר'}, "
        f"MAIL_USERNAME={'קיים' if mail_username else 'חסר'}, "
        f"MAIL_PASSWORD={'קיים' if mail_password else 'חסר'}"
    )

    if not mail_server or mail_server == "localhost":
        current_app.logger.error("[mail-test] SMTP לא מוגדר - MAIL_SERVER חסר או נשאר בברירת המחדל localhost")
        return jsonify({"success": False, "message": "SMTP לא מוגדר - יש להגדיר MAIL_SERVER ב-Render (Environment)."}), 500

    if not mail_username or not mail_password:
        current_app.logger.error("[mail-test] SMTP לא מוגדר במלואו - MAIL_USERNAME ו/או MAIL_PASSWORD חסרים")
        return jsonify({"success": False, "message": "פרטי התחברות לשרת המייל חסרים (MAIL_USERNAME / MAIL_PASSWORD)."}), 500

    # --- שלב 2: בניית ההודעה ---
    try:
        msg = Message("מייל בדיקה ממערכת המשימות", recipients=[user.email])
        msg.body = f"זהו מייל בדיקה. אם קיבלת אותו, הגדרות ה-SMTP (שרת: {mail_server}) תקינות ועובדות."
        current_app.logger.info(f"[mail-test] שלב בניית הודעה הושלם - יעד: user_id={user.id}")
    except Exception:
        current_app.logger.exception(f"[mail-test] שלב בניית הודעה נכשל - user_id={user.id}")
        return jsonify({"success": False, "message": "שגיאה פנימית בבניית ההודעה."}), 500

    # --- שלב 3: התחברות SMTP + שליחה ---
    import smtplib
    current_app.logger.info(f"[mail-test] שלב התחברות SMTP מתחיל - שרת={mail_server}, פורט={mail_port}")
    try:
        mail.send(msg)
    except smtplib.SMTPAuthenticationError:
        current_app.logger.exception("[mail-test] שלב שליחה נכשל - שגיאת הרשאה (אימות) מול שרת SMTP")
        return jsonify({"success": False, "message": "שגיאת הרשאה - שם המשתמש או הסיסמה לשרת המייל שגויים."}), 500
    except (smtplib.SMTPConnectError, smtplib.SMTPServerDisconnected, ConnectionRefusedError, TimeoutError, OSError):
        current_app.logger.exception(f"[mail-test] שלב התחברות SMTP נכשל - שרת={mail_server}, פורט={mail_port}")
        return jsonify({"success": False, "message": f"שגיאת התחברות לשרת המייל ({mail_server}:{mail_port}) - בדוק כתובת/פורט/חומת אש."}), 500
    except smtplib.SMTPException:
        current_app.logger.exception("[mail-test] שלב שליחה נכשל - שגיאת SMTP כללית")
        return jsonify({"success": False, "message": "שגיאת שרת המייל (SMTP) - בדוק את הגדרות התיבה."}), 500
    except Exception as e:
        current_app.logger.exception("[mail-test] שלב שליחה נכשל - שגיאה לא צפויה")
        return jsonify({"success": False, "message": f"שליחת המייל נכשלה: {type(e).__name__}"}), 500

    current_app.logger.info(f"[mail-test] שלב סיום - נשלח בהצלחה ל-user_id={user.id}")
    return jsonify({"success": True, "message": f"מייל בדיקה נשלח בהצלחה ל-{user.email}! (שרת: {mail_server})"})


@bp.route("/api/admin/mail/status")
@login_required
def mail_status():
    """
    אבחון בלי שליחת מייל בפועל - רק בודק אילו הגדרות SMTP קיימות.
    לא חושף סיסמה, שם משתמש מלא, או כל ערך רגיש אחר - רק true/false וברירת השרת/פורט.
    """
    if current_user.role != "admin":
        return jsonify({"success": False, "message": "אין הרשאה."}), 403

    mail_server = current_app.config.get("MAIL_SERVER")
    mail_port = current_app.config.get("MAIL_PORT")
    mail_username = current_app.config.get("MAIL_USERNAME")
    mail_password = current_app.config.get("MAIL_PASSWORD")

    smtp_configured = bool(mail_server and mail_server != "localhost" and mail_username and mail_password)

    return jsonify({
        "smtp_configured": smtp_configured,
        "server": mail_server,
        "port": mail_port,
        "username_exists": bool(mail_username),
    })


@bp.route("/api/notifications")
@login_required
def get_notifications():
    """מחזיר את ההתראות האחרונות של המשתמש + כמות שלא נקראו, לפעמון בסרגל העליון."""
    recent = Notification.query.filter_by(user_id=current_user.id).order_by(Notification.created_at.desc()).limit(15).all()
    unread_count = Notification.query.filter_by(user_id=current_user.id, is_read=False).count()
    return jsonify({
        "unread_count": unread_count,
        "notifications": [
            {
                "id": n.id,
                "message": n.message,
                "link": n.link,
                "icon": n.icon,
                "is_read": n.is_read,
                "created_at": n.created_at.strftime("%d/%m %H:%M") if n.created_at else "",
            }
            for n in recent
        ]
    })


@bp.route("/notifications/<int:id>/read", methods=["POST"])
@login_required
def mark_notification_read(id):
    n = Notification.query.filter_by(id=id, user_id=current_user.id).first()
    if n:
        n.is_read = True
        db.session.commit()
    return jsonify({"success": True})


@bp.route("/notifications/mark_all_read", methods=["POST"])
@login_required
def mark_all_notifications_read():
    Notification.query.filter_by(user_id=current_user.id, is_read=False).update({"is_read": True})
    db.session.commit()
    return jsonify({"success": True})


# =========================================================
# 🆘 SOS - התראת חירום רב-ערוצית (התראה פנימית + מייל אוטומטי + קישור וואטסאפ)
# =========================================================

def _sos_recipients(user):
    """
    קובע למי לשלוח SOS: קודם המנהל הישיר, ואם אין - מנהלי המחלקה, ואם אין אף אחד - כל מנהלי המערכת.
    מוודא שהשולח עצמו לעולם לא מקבל את ההתראה של עצמו.
    """
    recipients = []

    if user.direct_manager and user.direct_manager.id != user.id:
        recipients = [user.direct_manager]
    elif user.department_id:
        recipients = User.query.filter(
            User.department_id == user.department_id,
            User.role == "manager",
            User.id != user.id
        ).all()

    if not recipients:
        recipients = User.query.filter(User.role == "admin", User.id != user.id).all()

    return recipients


@bp.route("/sos", methods=["POST"])
@login_required
@limiter.limit("5 per hour")
def send_sos():
    message = request.form.get("message", "").strip() or "זקוק לעזרה דחופה!"
    recipients = _sos_recipients(current_user)

    full_message = f"🆘 SOS מ-{current_user.username}: {message}"

    email_sent_to, email_failed_to, whatsapp_targets = notify_recipients_multi_channel(
        recipients, full_message, link="/", icon="bi-exclamation-triangle-fill",
        email_subject=f"🆘 קריאת SOS מ-{current_user.username}",
    )

    if not recipients:
        return jsonify({"success": False, "message": "לא נמצא אף מנהל או אדמין לשלוח אליו התראה."}), 400

    return jsonify({
        "success": True,
        "recipient_count": len(recipients),
        "email_sent_to": email_sent_to,
        "email_failed_to": email_failed_to,
        "whatsapp_targets": whatsapp_targets,
    })


# =========================================================
# 📨 שליחת הודעה קבוצתית (למשתמשים נבחרים / מחלקה / כל הארגון)
# זמין רק ל-admin/manager, תמיד מוגבל להיקף ההרשאה של השולח (visible_user_ids)
# =========================================================

@bp.route("/send_bulk_message", methods=["POST"])
@login_required
@limiter.limit("20 per hour")
def send_bulk_message():
    if not can_create_tasks(current_user):
        return jsonify({"success": False, "message": "אין לך הרשאה לשלוח הודעות קבוצתיות."}), 403

    message = request.form.get("message", "").strip()
    if not message:
        return jsonify({"success": False, "message": "יש להזין תוכן להודעה."}), 400

    send_email = request.form.get("send_email") == "1"
    send_whatsapp = request.form.get("send_whatsapp") == "1"
    raw_ids = request.form.getlist("recipient_ids")

    # הגנה קריטית: לא משנה איזה מזהים נשלחו מהלקוח, אף אחד לא יכול לקבל הודעה
    # אם הוא לא בתוך היקף ההרשאה של השולח (visible_user_ids) - חוסם עקיפה של סינון בצד הלקוח.
    visible_ids = set(current_user.visible_user_ids())
    recipient_ids = [int(rid) for rid in raw_ids if rid.isdigit() and int(rid) in visible_ids and int(rid) != current_user.id]

    if not recipient_ids:
        return jsonify({"success": False, "message": "לא נבחרו נמענים תקינים בתוך היקף ההרשאה שלך."}), 400

    if not send_email and not send_whatsapp:
        return jsonify({"success": False, "message": "יש לבחור לפחות ערוץ שליחה אחד (מייל / וואטסאפ)."}), 400

    valid_recipients = [User.query.get(uid) for uid in recipient_ids]
    valid_recipients = [r for r in valid_recipients if r]
    notified_count = len(valid_recipients)

    email_sent_to, email_failed_to, whatsapp_targets = notify_recipients_multi_channel(
        valid_recipients, f"{current_user.username} שלח לך הודעה: {message[:120]}", link="/",
        icon="bi-chat-square-text", email_subject=f"הודעה ממערכת המשימות - {current_user.username}",
        send_email=send_email, send_whatsapp=send_whatsapp,
    )

    return jsonify({
        "success": True,
        "notified_count": notified_count,
        "email_sent_to": email_sent_to,
        "email_failed_to": email_failed_to,
        "whatsapp_targets": whatsapp_targets,
    })


# =========================================================
# 🔀 הקצאה קבוצתית - העברת כמה משימות נבחרות לאיש אחד בבת אחת
# =========================================================

@bp.route("/bulk_reassign_tasks", methods=["POST"])
@login_required
@limiter.limit("30 per hour")
def bulk_reassign_tasks():
    if not can_create_tasks(current_user):
        return jsonify({"success": False, "message": "אין לך הרשאה להקצות משימות."}), 403

    raw_task_ids = request.form.getlist("task_ids")
    new_assignee_id = request.form.get("assignee_id")

    if not new_assignee_id or not new_assignee_id.isdigit():
        return jsonify({"success": False, "message": "יש לבחור למי להקצות את המשימות."}), 400
    new_assignee_id = int(new_assignee_id)

    # הגנה: היעד להקצאה חייב להיות בתוך היקף ההרשאה של המשתמש (מחלקתו, או admin לכל אחד)
    if new_assignee_id not in current_user.visible_user_ids():
        return jsonify({"success": False, "message": "אינך רשאי להקצות משימות למשתמש זה."}), 403

    new_assignee = User.query.get(new_assignee_id)
    if not new_assignee:
        return jsonify({"success": False, "message": "המשתמש הנבחר לא נמצא."}), 400

    updated_count = 0
    skipped_count = 0

    for tid in raw_task_ids:
        if not tid.isdigit():
            continue
        task = Task.query.get(int(tid))
        # הגנה נוספת: כל משימה נבדקת בנפרד - לא ניתן לעקוף ולעדכן משימה שמחוץ להיקף ההרשאה
        if not task or not can_touch_task(current_user, task):
            skipped_count += 1
            continue

        if task.assigned_to_id != new_assignee_id:
            task.assigned_to_id = new_assignee_id
            db.session.commit()
            if new_assignee_id != current_user.id:
                notify_with_email(
                    new_assignee,
                    f"{current_user.username} הקצה לך משימה: \"{task.title}\"",
                    link=f"/edit/{task.id}", icon="bi-arrow-left-right",
                    email_subject=f"הוקצתה לך משימה: {task.title}",
                    email_body=(
                        f"שלום {new_assignee.username},\n\n"
                        f"{current_user.username} שייך אליך את המשימה \"{task.title}\" (הקצאה קבוצתית).\n"
                        f"{'תאריך יעד: ' + task.due_date.strftime('%d/%m/%Y') if task.due_date else ''}"
                    )
                )
        updated_count += 1

    if updated_count > 0:
        log_audit(
            current_user, "bulk_reassign_tasks", target_type="user", target_id=new_assignee_id,
            target_label=new_assignee.username,
            details=f"{updated_count} משימות הוקצו, {skipped_count} דולגו"
        )

    return jsonify({
        "success": True,
        "updated_count": updated_count,
        "skipped_count": skipped_count,
        "assignee_name": new_assignee.username,
    })


@bp.route("/bulk_delete_tasks", methods=["POST"])
@login_required
@limiter.limit("20 per hour")
def bulk_delete_tasks():
    """
    מחיקה קבוצתית - בכוונה מוגבלת ל-admin בלבד (לא manager), כדי שלא כל אחד
    יוכל למחוק כמות גדולה של משימות בבת אחת.
    """
    if current_user.role != "admin":
        return jsonify({"success": False, "message": "מחיקה קבוצתית זמינה רק למנהל מערכת."}), 403

    raw_task_ids = request.form.getlist("task_ids")
    deleted_count = 0
    deleted_titles = []

    for tid in raw_task_ids:
        if not tid.isdigit():
            continue
        task = Task.query.get(int(tid))
        if not task:
            continue
        deleted_titles.append(task.title)
        db.session.delete(task)
        deleted_count += 1

    db.session.commit()

    if deleted_count > 0:
        log_audit(
            current_user, "bulk_delete_tasks", target_type="task",
            details=f"{deleted_count} משימות נמחקו: {', '.join(deleted_titles[:10])}" + (" ..." if len(deleted_titles) > 10 else "")
        )

    return jsonify({"success": True, "deleted_count": deleted_count})


# =========================================================
# 📤 ייצוא דוחות (Excel / PDF) - מכבד את אותם הסינונים שהוצגו במסך (Phase 4 Updates)
# =========================================================

_STATUS_LABELS_HE = {"TODO": "לביצוע", "IN_PROGRESS": "בתהליך", "DONE": "בוצע"}
_PRIORITY_LABELS_HE = {"HIGH": "גבוהה", "MEDIUM": "בינונית", "LOW": "נמוכה"}


def _filtered_tasks_for_export():
    query = visible_task_query(current_user)
    query = apply_task_filters(query, current_user, request.args)
    return query.all()


@bp.route("/export/excel")
@login_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    tasks_list = _filtered_tasks_for_export()

    wb = Workbook()
    ws = wb.active
    ws.title = "משימות"
    ws.sheet_view.rightToLeft = True

    # הרחבת הכותרות (Phase 4)
    headers = ["כותרת", "תיאור", "סטטוס", "עדיפות", "אחראי", "תאריך יעד", "זמן מוערך (דק')", "משימת-אב", "מספר תלויות", "חזרתיות", "נוצר בתאריך"]
    ws.append(headers)
    header_fill = PatternFill(start_color="1D3A8A", end_color="1D3A8A", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for t in tasks_list:
        ws.append([
            t.title,
            t.description or "",
            _STATUS_LABELS_HE.get(t.status, t.status),
            _PRIORITY_LABELS_HE.get(t.priority, t.priority),
            t.assignee.username if t.assignee else "",
            t.due_date.strftime("%d/%m/%Y") if t.due_date else "",
            t.estimated_minutes or 0,
            t.parent_task.title if t.parent_task else "",
            t.dependencies.count(),
            RECURRENCE_LABELS_HE_EXPORT.get(t.recurrence, "חד פעמית"),
            t.created_at.strftime("%d/%m/%Y") if t.created_at else "",
        ])

    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 45)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"משימות_{date.today().strftime('%Y-%m-%d')}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/export/pdf")
@login_required
def export_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from bidi.algorithm import get_display

    # רושמים פונט שתומך בעברית (מצורף לפרויקט) - כדי לא להיות תלויים בפונטים שמותקנים על שרת Render
    font_dir = os.path.join(current_app.root_path, "static", "fonts")
    pdfmetrics.registerFont(TTFont("HebrewFont", os.path.join(font_dir, "DejaVuSans.ttf")))
    pdfmetrics.registerFont(TTFont("HebrewFont-Bold", os.path.join(font_dir, "DejaVuSans-Bold.ttf")))

    def he(text):
        """מכין טקסט עברי לתצוגה נכונה (RTL) בתוך PDF."""
        return get_display(str(text)) if text else ""

    tasks_list = _filtered_tasks_for_export()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1.5 * cm, bottomMargin=1.5 * cm)

    title_style = ParagraphStyle("HeTitle", fontName="HebrewFont-Bold", fontSize=18, alignment=1, spaceAfter=14)
    elements = [Paragraph(he(f"דוח משימות - {date.today().strftime('%d/%m/%Y')}"), title_style)]

    # הרחבת הכותרות (Phase 4)
    headers = [he("כותרת"), he("סטטוס"), he("עדיפות"), he("אחראי"), he("יעד"), he("זמן (דק')"), he("תלויות")]
    data = [headers]
    for t in tasks_list:
        data.append([
            he(t.title),
            he(_STATUS_LABELS_HE.get(t.status, t.status)),
            he(_PRIORITY_LABELS_HE.get(t.priority, t.priority)),
            he(t.assignee.username if t.assignee else "-"),
            t.due_date.strftime("%d/%m/%Y") if t.due_date else "-",
            str(t.estimated_minutes or 0),
            str(t.dependencies.count()),
        ])

    # התאמת רוחבי העמודות כך שייכנסו לרוחב דף A4 שוכב (~26 ס"מ נטו מרווח)
    table = Table(data, repeatRows=1, colWidths=[6.5 * cm, 2.5 * cm, 2.5 * cm, 3.5 * cm, 3 * cm, 2 * cm, 2 * cm])
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "HebrewFont"),
        ("FONTNAME", (0, 0), (-1, 0), "HebrewFont-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D3A8A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAF0FD")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    filename = f"tasks_report_{date.today().strftime('%Y-%m-%d')}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")


RECURRENCE_LABELS_HE_EXPORT = {"NONE": "חד פעמית", "DAILY": "כל יום", "WEEKLY": "כל שבוע", "MONTHLY": "כל חודש"}


# =========================================================
# ✉️ מערכת איפוס סיסמה במייל
# =========================================================

def send_reset_email(user):
    token = user.get_reset_token()
    reset_url = url_for('tasks.reset_token', token=token, _external=True)
    msg = Message('איפוס סיסמה', recipients=[user.email])  # sender נלקח אוטומטית מ-MAIL_DEFAULT_SENDER
    msg.body = f"לאיפוס לחץ על הקישור: {reset_url}"
    try:
        queue_email(msg)
    except Exception:
        current_app.logger.exception("הכנת מייל איפוס סיסמה נכשלה.")

@bp.route("/reset_password", methods=["GET", "POST"])
@limiter.limit("5 per hour", methods=["POST"])
def reset_request():
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        user = User.query.filter_by(email=email).first() if email else None
        if user:
            send_reset_email(user)
        flash("אם כתובת המייל קיימת במערכת ומשויכת לחשבון, נשלחו אליה הוראות לאיפוס הסיסמה.", "info")
        return redirect(url_for('tasks.login'))
    return render_template("reset_request.html")

@bp.route("/reset_password/<token>", methods=["GET", "POST"])
def reset_token(token):
    user = User.verify_reset_token(token)
    if not user:
        flash("קישור פג תוקף או שגוי.", "danger")
        return redirect(url_for('tasks.reset_request'))
    if request.method == "POST":
        new_password = request.form.get("password")
        is_strong, error_msg = validate_password_strength(new_password)
        if not is_strong:
            flash(error_msg, "danger")
            return render_template("reset_token.html")
        user.set_password(new_password)
        db.session.commit()
        flash("הסיסמה שונתה בהצלחה.", "success")
        return redirect(url_for('tasks.login'))
    return render_template("reset_token.html")


# =========================================================
# 🔐 פונקציות עזר משותפות - הרשאות גישה לכלי ניהול/תזכורות
# =========================================================

def _check_migration_key():
    if os.environ.get('ENABLE_ADMIN_TOOLS', '').lower() != 'true':
        return False
    secret = os.environ.get('MIGRATION_SECRET')
    if not secret:
        return False
    return request.args.get('key') == secret


def _check_reminder_key():
    """מפתח נפרד למנגנון התזכורות - לא תלוי בהפעלת כלי הניהול המסוכנים."""
    secret = os.environ.get('REMINDER_SECRET')
    if not secret:
        return False
    return request.args.get('key') == secret


# =========================================================
# ⏰ תזכורות מייל למשימות שמתקרב תאריך היעד שלהן
#
# הראוט הזה לא רץ לבד - הוא מיועד להיקרא פעם ביום ע"י שירות חיצוני
# (למשל cron-job.org, או GitHub Actions עם schedule) שקורא ל:
# https://<your-app>.onrender.com/api/send_due_reminders?key=<REMINDER_SECRET>
# =========================================================

@bp.route("/api/send_due_reminders")
def send_due_reminders():
    """
    שולח תזכורות (מייל + התראה פנימית) על משימות שמתקרב תאריך היעד שלהן.
    פרמטר ?when=today (ברירת מחדל) או ?when=tomorrow קובע אילו משימות נכללות.
    מיועד להיקרא ע"י cron חיצוני, למשל פעם ביום בבוקר:
    https://<your-app>.onrender.com/api/send_due_reminders?key=<REMINDER_SECRET>&when=today
    """
    if not _check_reminder_key():
        return "🔒 גישה חסומה. יש להגדיר REMINDER_SECRET ולספק ?key=... תואם.", 403

    when = request.args.get("when", "today")
    target_date = date.today() if when == "today" else date.today() + timedelta(days=1)
    when_label_he = "היום" if when == "today" else "מחר"

    due_tasks = Task.query.filter(
        Task.due_date == target_date,
        Task.status != "DONE",
        Task.assigned_to_id.isnot(None)
    ).all()

    sent, failed, notified = 0, 0, 0
    for task in due_tasks:
        if not task.assignee:
            continue

        # התראה פנימית - תמיד, בלי קשר אם יש מייל
        notify(
            task.assignee.id,
            f"תזכורת: המשימה \"{task.title}\" מתוכננת ל{when_label_he}",
            link=f"/edit/{task.id}",
            icon="bi-alarm"
        )
        notified += 1

        if not task.assignee.email:
            continue
        try:
            msg = Message(
                f"תזכורת: המשימה '{task.title}' מתוכננת ל{when_label_he}",
                recipients=[task.assignee.email]
            )
            msg.body = (
                f"שלום {task.assignee.username},\n\n"
                f"תזכורת שהמשימה \"{task.title}\" מתוכננת לתאריך {task.due_date.strftime('%d/%m/%Y')} ({when_label_he}).\n"
                f"{task.description or ''}\n\n"
                f"בהצלחה!"
            )
            queue_email(msg)
            sent += 1
        except Exception:
            current_app.logger.exception(f"הכנת תזכורת נכשלה עבור משימה {task.id}")
            failed += 1

    return jsonify({"when": when, "found": len(due_tasks), "notified": notified, "email_sent": sent, "email_failed": failed})


# =========================================================
# 🤖 הפעלת אוטומציות "task_overdue" - נועד להיקרא ע"י cron חיצוני פעם ביום
# https://<your-app>.onrender.com/api/run_overdue_automations?key=<REMINDER_SECRET>
# =========================================================

@bp.route("/api/run_overdue_automations")
def run_overdue_automations():
    if not _check_reminder_key():
        return "🔒 גישה חסומה. יש להגדיר REMINDER_SECRET ולספק ?key=... תואם.", 403

    overdue_tasks = Task.query.filter(
        Task.due_date < date.today(),
        Task.status != "DONE"
    ).all()

    for task in overdue_tasks:
        automation_service.trigger("task_overdue", task)

    return jsonify({"success": True, "checked": len(overdue_tasks)})


# =========================================================
# 📈 תמונת מצב יומית לצורך גרף מגמות (השלמת משימות לאורך זמן)
#
# מיועד להיקרא ע"י cron חיצוני פעם ביום (לרוב בסוף היום):
# https://<your-app>.onrender.com/api/snapshot_daily_stats?key=<REMINDER_SECRET>
# אידמפוטנטי - קריאה נוספת באותו יום מעדכנת את השורה הקיימת במקום ליצור כפילות.
# =========================================================

@bp.route("/api/snapshot_daily_stats")
def snapshot_daily_stats():
    if not _check_reminder_key():
        return "🔒 גישה חסומה. יש להגדיר REMINDER_SECRET ולספק ?key=... תואם.", 403

    from app.models.daily_stat import DailyStat

    today_date = date.today()
    all_tasks = Task.query.all()

    def upsert_stat(department_id, tasks_subset):
        existing = DailyStat.query.filter_by(stat_date=today_date, department_id=department_id).first()
        total = len(tasks_subset)
        done = len([t for t in tasks_subset if t.status == "DONE"])
        if existing:
            existing.total_tasks = total
            existing.done_tasks = done
        else:
            db.session.add(DailyStat(stat_date=today_date, department_id=department_id, total_tasks=total, done_tasks=done))

    # תמונת מצב כלל-ארגונית
    upsert_stat(None, all_tasks)

    # תמונת מצב לכל מחלקה בנפרד
    users_by_dept = {}
    for u in User.query.all():
        users_by_dept.setdefault(u.department_id, []).append(u.id)

    for dept in Department.query.all():
        member_ids = users_by_dept.get(dept.id, [])
        dept_tasks = [t for t in all_tasks if t.assigned_to_id in member_ids]
        upsert_stat(dept.id, dept_tasks)

    db.session.commit()
    return jsonify({"success": True, "date": today_date.isoformat(), "total_tasks": len(all_tasks)})


@bp.route("/api/stats_trend")
@login_required
def stats_trend():
    """מחזיר את 30 הימים האחרונים של תמונות מצב, לצורך גרף המגמות בדשבורד (admin בלבד)."""
    if current_user.role != "admin":
        return jsonify({"success": False, "message": "אין הרשאה."}), 403

    from app.models.daily_stat import DailyStat

    cutoff = date.today() - timedelta(days=30)
    stats = DailyStat.query.filter(
        DailyStat.department_id.is_(None),
        DailyStat.stat_date >= cutoff
    ).order_by(DailyStat.stat_date.asc()).all()

    return jsonify({
        "labels": [s.stat_date.strftime("%d/%m") for s in stats],
        "completion_percent": [s.completion_percent for s in stats],
        "total_tasks": [s.total_tasks for s in stats],
    })


# =========================================================
# 🛠️ כלי חילוץ ושדרוג אוטומטי למסד הנתונים
#
# ⚠️ הראוטים האלה מבצעים שינויים במסד הנתונים ולכן מוגנים ב-"מפתח מיגרציה".
# יש להגדיר משתני סביבה ENABLE_ADMIN_TOOLS=true ו-MIGRATION_SECRET (ב-Render)
# ולהעביר את המפתח כפרמטר ?key=... אחרת הראוטים חסומים לגמרי מטעמי אבטחה.
# =========================================================

@bp.route("/fix-db")
def fix_db():
    if not _check_migration_key():
        return "🔒 גישה חסומה. יש להגדיר ENABLE_ADMIN_TOOLS=true ו-MIGRATION_SECRET, ולספק ?key=... תואם.", 403
    log_audit(None, "use_fix_db_tool")

    output = []
    try:
        db.session.execute(text("ALTER TABLE \"user\" ADD COLUMN role VARCHAR(20) DEFAULT 'employee';"))
        db.session.commit()
        output.append("✅ עמודת 'role' נוספה לטבלת המשתמשים.")
    except Exception:
        db.session.rollback()
        output.append("ℹ️ עמודת 'role' כבר קיימת.")

    try:
        db.session.execute(text('ALTER TABLE task ADD COLUMN assigned_to_id INTEGER;'))
        db.session.commit()
        output.append("✅ עמודת 'assigned_to_id' נוספה לטבלת המשימות.")
        db.session.execute(text('UPDATE task SET assigned_to_id = user_id WHERE assigned_to_id IS NULL;'))
        db.session.commit()
    except Exception:
        db.session.rollback()
        output.append("ℹ️ עמודת 'assigned_to_id' כבר קיימת.")

    try:
        db.session.execute(text("UPDATE \"user\" SET role = 'manager' WHERE username = 'mv';"))
        db.session.commit()
        output.append("👑 המשתמש 'mv' הוגדר בהצלחה כמנהל המערכת.")
    except Exception:
        db.session.rollback()

    return "<br>".join(output) + "<br><br><b>השדרוג למערכת ארגונית הסתיים! חזור לאתר והתחבר.</b>"


@bp.route("/upgrade-permissions")
def upgrade_permissions():
    """
    מיגרציה חד-פעמית לשדרוג מודל ההרשאות:
    - יוצר טבלת department (אם לא קיימת).
    - מוסיף לטבלת user את department_id ו-manager_id (אם חסרים).
    - הופך את המשתמש 'mv' (ומשתמשי 'manager' ישנים) ל-role='admin'.
    יש לקרוא לראוט הזה פעם אחת בלבד, עם ?key=<MIGRATION_SECRET>.
    """
    if not _check_migration_key():
        return "🔒 גישה חסומה. יש להגדיר ENABLE_ADMIN_TOOLS=true ו-MIGRATION_SECRET, ולספק ?key=... תואם.", 403
    log_audit(None, "use_upgrade_permissions_tool")

    output = []

    # 1. יצירת כל הטבלאות החדשות שעדיין לא קיימות (כולל department)
    try:
        db.create_all()
        output.append("✅ טבלת 'department' נוצרה/קיימת.")
    except Exception as e:
        output.append(f"⚠️ שגיאה ביצירת טבלאות: {e}")

    # 2. הוספת עמודות חדשות לטבלת user
    for column_sql, label in [
        ('ALTER TABLE "user" ADD COLUMN department_id INTEGER;', "department_id"),
        ('ALTER TABLE "user" ADD COLUMN manager_id INTEGER;', "manager_id"),
    ]:
        try:
            db.session.execute(text(column_sql))
            db.session.commit()
            output.append(f"✅ עמודת '{label}' נוספה לטבלת המשתמשים.")
        except Exception:
            db.session.rollback()
            output.append(f"ℹ️ עמודת '{label}' כבר קיימת.")

    # 3. הפיכת מנהל המערכת הראשי ל-admin (במקום 'manager' הישן)
    try:
        result = db.session.execute(text("UPDATE \"user\" SET role = 'admin' WHERE username = 'mv';"))
        db.session.commit()
        output.append("👑 המשתמש 'mv' הוגדר כמנהל מערכת (admin).")
    except Exception as e:
        db.session.rollback()
        output.append(f"⚠️ שגיאה בעדכון 'mv': {e}")

    output.append(
        "<br><b>שלב הבא:</b> היכנס לפאנל הניהול (/admin) → נהל מחלקות → "
        "ואז שייך כל 'מנהל תחום' למחלקה המתאימה דרך עריכת משתמש."
    )

    return "<br>".join(output)

@bp.route("/rescue")
def rescue():
    if not _check_migration_key():
        return "🔒 גישה חסומה. יש להגדיר ENABLE_ADMIN_TOOLS=true ו-MIGRATION_SECRET, ולספק ?key=... תואם.", 403
    log_audit(None, "use_rescue_tool", details="איפוס סיסמת mv ל-123456 דרך /rescue")
    try:
        db.create_all()
        admin = User.query.filter_by(username='mv').first()
        if admin:
            admin.set_password("123456")
            admin.must_change_password = True
            db.session.commit()
            return f"✅ הסיסמה אופסה ל-123456 עבור {admin.email}. תידרש להחליף אותה בכניסה הבאה."
        else:
            admin = User(username='mv', email='admin@test.com')
            admin.set_password("123456")
            admin.must_change_password = True
            db.session.add(admin)
            db.session.commit()
            return "✅ יוזר מנהל mv נוצר עם הסיסמה 123456. תידרש להחליף אותה בכניסה הבאה."
    except Exception as e:
        db.session.rollback()
        return f"❌ שגיאה: {e}"
